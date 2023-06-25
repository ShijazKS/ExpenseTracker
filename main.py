from tkinter  import * 
from datetime import date
from tkinter import messagebox
import tkinter
from tkinter import filedialog
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference


#from screen import *



background = "#06283D"
topbg = "#F6283D"
ms = ['jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec']


root = Tk()
root.title("Expense Tracker System")
root.geometry("720x420")
root.resizable(False,False)
root.config(bg=background)


# excel_sheet
file =pathlib.Path('My_Expense.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Date"
    sheet['B1'] = "Month"
    sheet['C1'] = "Snack"
    sheet['D1'] = "Morning"
    sheet['E1'] = "Noon"
    sheet['F1'] = "Night"
    sheet['G1'] = "Total"


    file.save('My_Expense.xlsx')

####chart in excel#########


###########################

###############SAVE#########################
def Save():
    D0 = Date.get()
    M1 = Month.get()
    if M1 == 'select':
        messagebox.showerror("error","Select Month")
        return False
    else:    
        S1 = Snack.get()
        B1 = bfast.get()
        L1 = lunch.get()
        D1 = dinner.get()
        T1 = int(S1)+int(B1)+int(L1)+int(D1)

        print(M1)
    file = openpyxl.load_workbook('My_Expense.xlsx')
    sheet = file.active

    sheet.cell(column=1,row=sheet.max_row+1,value=D0)
    sheet.cell(column=2,row=sheet.max_row,value=M1)
    sheet.cell(column=3,row=sheet.max_row,value=S1)
    sheet.cell(column=4,row=sheet.max_row,value=B1)
    sheet.cell(column=5,row=sheet.max_row,value=L1)
    sheet.cell(column=6,row=sheet.max_row,value=D1)
    sheet.cell(column=7,row=sheet.max_row,value=T1)



    file.save(r'My_Expense.xlsx')

    

    return True

###########################################


def submit_expense():
    # Get the expense details
    date_expense = date_entry.get()
    month_expense = month_entry.get()
    snack_expense = snack_entry.get()
    bfast_expense = bfast_entry.get()
    lunch_expense = lunch_entry.get()
    dinner_expense = dinner_entry.get()


     # Validate the input
    if not bfast_expense.isdigit() or not lunch_expense.isdigit() or not dinner_expense.isdigit():
        messagebox.showerror("Invalid Input", "Please enter numeric values for expenses.")
        return

    mCheck = Save()
    
    if mCheck == True:

        total = int(bfast_expense) + int(lunch_expense) + int(dinner_expense) + int(snack_expense)
        result = "Total is : " + str(total)

        # Display success message
        messagebox.showinfo("Expense Tracker",result)


    # Clear the entry fields
    bfast_entry.delete(0,END)
    lunch_entry.delete(0,END)
    dinner_entry.delete(0,END)

    bfast_entry.insert(0,'0')
    lunch_entry.insert(0,'0')
    dinner_entry.insert(0,'0')
    


#top frames
Label(root,text="EXPENSE TRACKER",font='arial 20 bold',width=10,height=2,bg=topbg,fg="#fff", anchor='center').pack(side=TOP,expand=False,fill=X)

#############
#Date
Label(root,text="Date:",font='arial 13',fg='#fff',bg=background).place(x=15,y=125)
Date = StringVar()
today = date.today()
d1= today.strftime("%d/%m/%Y")
date_entry = Entry(root,textvariable=Date,width=11,font='arial 10')
date_entry.place(x=63,y=127)
Date.set(d1)

#Month
Label(root,text="Month:",font='arial 13',fg='#fff',bg=background).place(x=265,y=125)
Month = StringVar()
month_entry = Combobox(root,textvariable=Month,values=ms,font="Roboto 10",width=7,state='r')
month_entry.set("select")
month_entry.place(x=323,y=127)



#Week
Label(root,text="Snack:",font='arial 13',fg='#fff',bg=background).place(x=550,y=125)
Snack = StringVar()
snack_entry = Entry(root,textvariable=Snack,width=6,font='arial 10')
snack_entry.place(x=613,y=127)
snack_entry.insert(0,'0')

#########

#############
#Bfast
Label(root,text="Breakfast:",font='arial 13',fg='#fff',bg=background).place(x=15,y=195)
bfast = StringVar()
bfast_entry = Entry(root,width=6,textvariable=bfast,font='arial 10')
bfast_entry.place(x=93,y=197)
bfast_entry.insert(0,'0')


#Lunch
Label(root,text="Lunch:",font='arial 13',fg='#fff',bg=background).place(x=265,y=195)
lunch = StringVar()
lunch_entry = Entry(root,textvariable=lunch,width=6,font='arial 10')
lunch_entry.place(x=323,y=197)
lunch_entry.insert(0,'0')


#Dinner
Label(root,text="Dinner:",font='arial 13',fg='#fff',bg=background).place(x=550,y=195)
dinner = StringVar()
dinner_entry = Entry(root,textvariable=dinner,width=6,font='arial 10')
dinner_entry.place(x=613,y=197)
dinner_entry.insert(0,'0')


submit_button = Button(root, text="Submit",command=submit_expense,bg='lightblue',width=13,font='arial 10 bold')
submit_button.place(x=294,y=257)

#########
root.mainloop()
